import pandas as pd 
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font,colors,Color,Alignment,PatternFill,GradientFill,Border,Side
from openpyxl.styles import NamedStyle

df1=pd.read_excel('shifts.xlsx', sheet_name='Sheet')
df2=pd.read_excel('shifts.xlsx', sheet_name='Sheet1')
df3=pd.read_excel('shift_3.xlsx')

dfall=pd.concat([df1,df2,df3],sort=False)
#print(dfall) # prints all 3 sheets val
#print(dfall.loc[50]) # print 50th row in all 3 sheets
#print(dfall.groupby(['Shift']).mean()['Units sold'])

toexcel=dfall.to_excel('allshifts.xlsx',index=None)

wb=load_workbook('allshifts.xlsx')
ws=wb.active

totalcol=ws['G1']
totalcol.font=Font(bold=True)
totalcol.value='Total'

e_col,f_col=['E','F']

for row in range(2,300):
    result_cell = 'G{}'.format(row)
    e_value=ws[e_col + str(row)].value
    f_value=ws[f_col + str(row)].value

    ws[result_cell]=e_value * f_value

    wb.save('total.xlsx')