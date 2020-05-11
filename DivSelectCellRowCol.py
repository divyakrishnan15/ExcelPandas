from openpyxl.workbook import Workbook
from openpyxl import load_workbook

wb=load_workbook('regions.xlsx')
ws=wb.active

cellrange=ws['A1':'C1']
#print(cellrange)
colrange=ws['A':'C']
#print(colrange)

rowrange=ws[1:5]

for row in ws.iter_rows(min_row=1,max_row=2,max_col=3,values_only=True):
    for cell in row:
        print(cell)